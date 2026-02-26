from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def parse_timestamp(value: str) -> datetime:
    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(
        "Unsupported timestamp format. Use one of: YYYY-MM-DD HH:MM[:SS], "
        "YYYY/MM/DD HH:MM[:SS], or ISO format YYYY-MM-DDTHH:MM[:SS]."
    )


def read_rows(csv_path: Path) -> list[dict[str, str]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        required = {"timestamp", "speaker", "message"}
        headers = set(h.lower() for h in (reader.fieldnames or []))
        if not required.issubset(headers):
            raise ValueError(
                "CSV must include headers: timestamp,speaker,message"
            )

        normalized: list[dict[str, str]] = []
        for row in reader:
            lowered = {k.lower(): (v or "") for k, v in row.items()}
            normalized.append(
                {
                    "timestamp": lowered["timestamp"],
                    "speaker": lowered["speaker"],
                    "message": lowered["message"],
                }
            )
        return normalized


def group_by_hour(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    buckets: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        ts = parse_timestamp(row["timestamp"])
        key = ts.strftime("%Y-%m-%d_%H")
        buckets[key].append(
            {
                "timestamp": ts.strftime("%Y-%m-%d %H:%M:%S"),
                "speaker": row["speaker"],
                "message": row["message"],
            }
        )
    return dict(sorted(buckets.items(), key=lambda item: item[0]))


def autosize_columns(sheet) -> None:
    widths = {"A": 19, "B": 14, "C": 60}
    for col, default in widths.items():
        max_len = default
        for cell in sheet[col]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = min(len(value), 120)
        sheet.column_dimensions[col].width = max_len + 2


def write_workbook(grouped: dict[str, list[dict[str, str]]], output_path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for hour_key, items in grouped.items():
        sheet_name = hour_key[:31]
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(["Timestamp", "Speaker", "Message"])
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for item in items:
            sheet.append([item["timestamp"], item["speaker"], item["message"]])
        autosize_columns(sheet)

    if not grouped:
        sheet = workbook.create_sheet(title="NoData")
        sheet.append(["Timestamp", "Speaker", "Message"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert chat CSV into Excel with one sheet per hour."
    )
    parser.add_argument(
        "--input",
        default="chat_logs/chat_history.csv",
        help="Input CSV file with headers: timestamp,speaker,message",
    )
    parser.add_argument(
        "--output",
        default="chat_logs/chat_history_hourly.xlsx",
        help="Output Excel (.xlsx) file path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    rows = read_rows(input_path)
    grouped = group_by_hour(rows)
    write_workbook(grouped, output_path)
    print(f"Saved: {output_path} (sheets: {len(grouped)})")


if __name__ == "__main__":
    main()
